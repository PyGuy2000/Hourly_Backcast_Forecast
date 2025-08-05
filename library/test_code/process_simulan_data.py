import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
input_file = r'C:\Users\kaczanor\OneDrive - Enbridge Inc\Documents\Python\EDC Hourly Capacity Factor Q2 2024\inputs\Simulan\Q3_2024_Similan_DataFile_Base.xlsx'  # Replace with your actual file path
sheet_name = 'Hourly'

# Read the data from the specified sheet
df = pd.read_excel(input_file, sheet_name=sheet_name)

# xl = pd.ExcelFile(input_file)
# df = xl.parse(xl.sheet_names[0])  # Assuming the data is in the first sheet

# Get unique years from the Date column
df['Date'] = pd.to_datetime(df['Date'])
df['Year'] = df['Date'].dt.year
years = df['Year'].unique()

# Load the workbook
wb = load_workbook(input_file)

# Create new worksheets for each year and filter data
for year in years:
    # Filter data for the specific year
    df_year = df[df['Year'] == year]
    
    # Drop the 'Year' column as it's no longer needed
    df_year = df_year.drop(columns=['Year'])
    
    # Create a new worksheet for the year
    sheet_name = str(year)
    ws = wb.create_sheet(title=sheet_name)
    
    # Write the dataframe to the new worksheet
    for r_idx, row in enumerate(df_year.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)
    
    # Write the column headers
    for c_idx, col in enumerate(df_year.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)

# Save the workbook with the new sheets
wb.save(input_file)