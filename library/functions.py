from library import globals as gbl
from library import config as cfg
from library import initializer as init # pylint: disable=import-error
from library.class_objects.other_classes.classes import TrackedDict, ForecastFile

import os
import sys
import threading
from queue import Queue # needed to return wb object from load_workbook_thread() function
import time
from tqdm import tqdm
import glob
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import matplotlib.ticker as ticker  # Import the ticker module
import plotly.io as pio # pylint: disable=no-member
import plotly.graph_objects as go # pylint: disable=no-member
import pandas as pd
import numpy as np
import re
import csv 
from prettytable import PrettyTable
import pprint as pp
from tabulate import tabulate
import openpyxl.utils.cell as cell_utils
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import shutil



########################################
def save_data_as_csv_print_data_table_and_update_dictdate_dict(
            output_path, 
            gen_type, 
            data_dict_filename_str,
            json_dict, 
            json_dict_path, 
            gen_data
        ):

    # Create output path
    directory_path_str = os.path.join(output_path, gen_type)
    # Check if the directory exists, if not, create it.
    if not os.path.exists(directory_path_str):
        os.makedirs(directory_path_str)
    # Add file name to path
    output_file = os.path.join(directory_path_str, data_dict_filename_str)
    
    key, headers, table = print_lcoe_table({gen_type: gen_data})

    # Save to csv file
    save_table_to_csv(key, headers, table, output_file)
    
    # Save to json file
    json_dict.save_to_json(json_dict_path)

    return output_file


def print_dict_as_table(d):
    # Create a PrettyTable object
    table = PrettyTable()
    
    # Add columns to the table
    table.field_names = ["Key", "Value"]
    
    # Add rows to the table
    for key, value in d.items():
        table.add_row([key, value])
    
    # Print the table
    print(table)
########################################
# Create multiple stacked bar for lcoe for all generators
# using matplotlib. Do not use the 'Total LCOE' item for this chart
# as it will be included in the total LCOE chart. Format the y axis as LCOE in $0.00 format.
# and the x axis with genertor names. The chart should have a legend with the generator names.
def stacked_lcoe_chart_all_generators(lcoe_data_by_generator):

    # Create a matplotlib figure
    fig, ax = plt.subplots(figsize=(12, 8))

    # Get the list of categories from the first generator (assuming all have the same categories)
    first_key = next(iter(lcoe_data_by_generator))
    categories = list(lcoe_data_by_generator[first_key][0]['categories'].keys())
    categories.remove('Total LCOE')  # We don't want to plot the 'Total LCOE' category in 'lcoe_data_by_generator' as it will become a stack on top of the other categories

    # for generator_key, lcoe_data_list in lcoe_data_by_generator.items():
    #     print(f"\nLCOE Data for {generator_key}:\n")
    #     for entry in lcoe_data_list:
    #         categories = entry['categories'].keys()
    #         print(f"categories: {categories}")
    #     print("-" *40)

    # # We don't want to plot the 'Total LCOE' category in 'lcoe_data_by_generator' as it will become a stack on top of the other categories
    # keys_to_remove = ['Total LCOE']
    # for key in keys_to_remove:
    #     if key in categories:
    #         del lcoe_data_by_generator(key)

    # Print categories
    print(f"categories: {categories}")
    pp.pprint(lcoe_data_by_generator)

     # Define a color map for the categories
    colors = plt.cm.get_cmap('gist_ncar', len(categories))
    #colors = plt.cm.get_cmap('tab20')
    #color_list = colors.colors
    # Map colors to categories
    category_colors = {category: colors(i) for i, category in enumerate(categories)}
    #category_colors = {category: color_list[i % len(color_list)] for i, category in enumerate(categories)}
    

    # Extract total LCOE for each generator from the provided data
    total_lcoe_by_generator = {
        key: lcoe_data[0]['categories']['Total LCOE']
        for key, lcoe_data in lcoe_data_by_generator.items()
    }

    # Sort generators by total LCOE in descending order
    sorted_generators = sorted(total_lcoe_by_generator, key=total_lcoe_by_generator.get, reverse=True)
    print(f"sorted_generators: {sorted_generators}")


    # Extract capacity factors and prepare new x-axis labels
    x_labels = []

    for key in sorted_generators:
        lcoe_data = lcoe_data_by_generator[key]
        capacity_factors = [entry['capacity_factor'] for entry in lcoe_data]
        avg_capacity_factor = np.mean(capacity_factors) * 100  # Assuming you want the average capacity factor in percentage
        x_labels.append(f"{key}\n@ {avg_capacity_factor:.1f}% CF")

    # Initialize a dictionary to hold bottom values for each generator
    bottoms = {key: np.zeros(len(lcoe_data)) for key, lcoe_data in lcoe_data_by_generator.items()}

    # Loop through all categories and sorted generators to plot
    for category in categories:
        for idx, key in enumerate(sorted_generators):
            lcoe_data = lcoe_data_by_generator[key]
            values = [entry['categories'][category] for entry in lcoe_data]
            ax.bar(idx, values, bottom=bottoms[key], label=category if idx == 0 else "", color=category_colors[category])
            bottoms[key] += np.array(values)

    # Set the labels and title
    ax.set_xlabel('Generators')
    ax.set_ylabel('LCOE ($/MWh)')
    ax.set_title('LCOE by Category for All Generators')

    # Set the new x-axis labels
    ax.set_xticks(range(len(x_labels)))
    ax.set_xticklabels(x_labels, rotation=45, ha="right")

    # Create a single legend
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys())

    # Format the y-axis labels as currency
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(currency_format))

    # Show the plot
    chart_title = 'LCOE by Category for All Generators'
    filename = sanitize_filename(chart_title)
    display_figure(fig, f"{filename}.png")
########################################
def save_figure(fig, filepath):
    print(f" filepath: {filepath}")
    """
    Function to save the figure to a file.
    
    Parameters:
    fig: The figure object to be saved.
    filepath: The full path to save the figure.
    Helper function to save the figure to the specified path.
    """

    try:
        if isinstance(fig, plt.Figure):
            fig.savefig(filepath, bbox_inches='tight')
        elif isinstance(fig, go.Figure):
            fig.write_image(filepath)
        else:
            raise TypeError("Unsupported figure type. Only Matplotlib and Plotly figures are supported.")
    except Exception as e:
        logging.error(f"Failed to save figure: {filepath}", exc_info=True)
        print(f"An error occurred while saving the figure: {e}")

    print(f"Figure saved as {filepath}")
############################################
def display_figure(fig, filename=None, subfolder_name=''):
    """
    Function to save the chart based on the value of SAVE_VISUALIZATIONS
    and display the chart based on the value of DISPLAY_VISUALIZATIONS.
    
    Parameters:
    fig: The figure object to be displayed.
    filename: The filename to save the figure. If None, the figure will not be saved.
    subfolder_name: Optional subfolder name for saving the figure.
    """
    

    if cfg.SAVE_VISUALIZATIONS and filename:
        # Create the full path using the global directory variables
        gbl_root_dir = 'C:/Users/kaczanor/OneDrive - Enbridge Inc/Documents/Python/EDC Hourly Capacity Factor Q2 2024'
        gbl_output_dir = "outputs/"
        gbl_image_folder = 'image_data'

        
        output_directory = os.path.join(gbl_root_dir, gbl_output_dir, gbl_image_folder)
        full_path = create_path(output_directory, filename, subfolder_name)
        print(f" full_path: {full_path}")

        save_figure(fig, full_path)

    if cfg.DISPLAY_VISUALIZATIONS:
        if isinstance(fig, plt.Figure):
            plt.show()
        elif isinstance(fig, go.Figure):
            fig.show()
        else:
            raise TypeError("Unsupported figure type. Only Matplotlib and Plotly figures are supported.")
############################################
# def create_excel_output_table(
#             run_scenario,
#             run_case,
#             excel_template_path,
#             output_excel_folder,
#             csv_files,
#             stochastic_seeds_used
# ):

#     # Load the Excel template
#     workbook = load_workbook(excel_template_path)
#     worksheet = workbook['Monthly Py Frcst 1']

#     # Adjust the CSV file ranges based on the number of seeds used
#     adjusted_csv_files = adjust_excel_ranges(csv_files, stochastic_seeds_used)

#     # Loop through CSV files and paste data into Excel
#     for csv_file, cell_range in adjusted_csv_files.items():
#         paste_csv_to_excel(csv_file, cell_range, worksheet, csv_folder)

#     # Define the output file name based on run_case and run_scenario
#     output_file_var = f"{run_case}_{run_scenario}.xlsx"
#     output_excel_path = os.path.join(output_excel_folder, output_file_var)

#     # Save the updated Excel file with the new name
#     workbook.save(output_excel_path)

#     return
######################################
# The sanitize_filename function replaces spaces with underscores and removes special characters that are not allowed in filenames.
# This function is applied to both the filename and subfolder_name inside the create_path function.
def sanitize_filename(filename):
    # Replace spaces with underscores
    filename = filename.replace(' ', '_')
    
    # Replace scientific notation with readable format
    filename = re.sub(r'(\d+)\.0+', r'\1', filename)
    
    # Remove or replace any special characters that are not allowed in filenames
    filename = re.sub(r'[<>:"/\\|?*@%]', '', filename)
    
    return filename   
######################################
def create_path(input_dir, filename, subfolder_name=''):
    # Sanitize the subfolder_name and filename
    sanitized_subfolder_name = sanitize_filename(subfolder_name)
    sanitized_filename = sanitize_filename(filename)
    
    # Use os.path.join for all components
    full_path = os.path.normpath(os.path.join(input_dir, sanitized_subfolder_name, sanitized_filename))
    return full_path
######################################
def load_csv_with_datetime_index(file_path):
    """
    Loads a CSV file into a DataFrame and sets the index to a datetime index.

    Parameters:
    file_path (str): The path to the CSV file.

    Returns:
    pd.DataFrame: The DataFrame with a datetime index.
    """
    # Read the CSV file, parse the 'DateTime' column as dates
    df = pd.read_csv(file_path, parse_dates=['DateTime'])
    
    # Ensure that the 'DateTime' column is set as the index
    df.set_index('DateTime', inplace=True)
    
    return df

def load_workbook_thread(input_file, data_only, progress_event, queue,sheet_name=None):
    wb = load_workbook(input_file, keep_links=False, data_only=data_only)
    wb.calculation.calculationMode = "manual"
    #queue.put(wb)  # Put the workbook object into the queue
    #progress_event.set()  # Signal that the workbook has been loaded
    
    # Edited to account for worksheet designation
    if sheet_name:
        ws = wb[sheet_name]  # Load the specified sheet
    else:
        ws = wb.active  # Default to the first sheet if no sheet name is provided
    queue.put((wb, ws))  # Put the workbook and worksheet objects into the queue
    progress_event.set()  # Signal that the workbook has been loaded
    return wb

def show_progress_bar(progress_event, description="Loading workbook"):
    with tqdm(desc=description, total=100, ncols=100) as pbar:
        while not progress_event.is_set():
            pbar.update(1)
            time.sleep(0.1)
            pbar.n = pbar.n % 100  # Reset progress bar to loop
            
##############################
def process_excel_data(frcst_input_excel_file_path, ws=None):
    # Set up the logger
    logging.basicConfig(level=logging.INFO)

    print("Loading Excel File")

    progress_event = threading.Event()
    # Create a queue to receive the workbook object
    queue = Queue()  # needed to return wb object from load_workbook_thread() function
    # Note 'load_workbook_thread()' is a function in the functions.py file
    thread = threading.Thread(target=load_workbook_thread, args=(frcst_input_excel_file_path, True, progress_event, queue, ws))
    thread.start()
    # Show the progress bar
    show_progress_bar(progress_event)
    # Wait for the thread to finish
    thread.join()
    # Retrieve the workbook and worksheet objects from the queue
    wb, ws = queue.get()
    print("Workbook loaded successfully!")
    return wb, ws
            
#########################################
def create_file_from_string(filename_tuple):
    filename_str, extension = filename_tuple

    # Define allowed extensions
    allowed_extensions = ['csv', 'xlsx', 'png', 'json']
    if extension not in allowed_extensions:
        raise ValueError(f"Extension {extension} is not supported. Allowed extensions are {allowed_extensions}.")

    # Check for problematic characters in the filename
    if not re.match(r'^[\w\-]+$', filename_str):
        raise ValueError("Filename contains problematic characters. Only alphanumeric characters, underscores, and hyphens are allowed.")

    # Create the filename with the extension
    full_filename = f"{filename_str}.{extension}"
    
    # Write to the file (for demonstration purposes, we'll just create an empty file)
    with open(full_filename, 'w') as file:
        file.write("")

    return full_filename
#########################################
def adjust_excel_ranges(csv_files, stochastic_seeds_used):
    adjusted_csv_files = {}
    for filename, range_str in csv_files.items():
        start_cell, end_cell = range_str.split(":")
        start_col, start_row = cell_utils.coordinate_from_string(start_cell)
        end_col, end_row = cell_utils.coordinate_from_string(end_cell)
        
        new_end_col = cell_utils.get_column_letter(cell_utils.column_index_from_string(start_col) + stochastic_seeds_used - 1)
        new_range = f"{start_col}{start_row}:{new_end_col}{end_row}"
        
        adjusted_csv_files[filename] = new_range
    
    return adjusted_csv_files
#########################################
# Function to paste CSV data into the specified cell range in the worksheet
def paste_csv_to_excel(csv_file, cell_range, worksheet, csv_folder):
    # Read the CSV file, skipping the header row
    df = pd.read_csv(os.path.join(csv_folder, csv_file), skiprows=1)
    print(f"df: {df}")

    # Split the cell range into start and end cells
    start_cell, end_cell = cell_range.split(':')
    print(f"start_cell: {start_cell}, end_cell: {end_cell}")
    start_col, start_row = start_cell[0], int(start_cell[1:])
    print(f"start_col: {start_col}, start_row: {start_row}")
    end_col, end_row = end_cell[0], int(end_cell[1:])
    print(f"end_col: {end_col}, end_row: {end_row}")

    # Iterate over the DataFrame and write data to the worksheet
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            worksheet.cell(row=start_row + row_idx, column=8 + col_idx, value=value)
###############
def currency_format(x, pos):
    """The two args are the value and tick position."""
    return f'${x:,.2f}'
###############

def graph_lcoe(lcoe_data_by_generator):
     

    fig = plt.figure(figsize=(10, 6))
    
    for key, lcoe_data in lcoe_data_by_generator.items():
        print(f"\nLCOE Data for {key}:\n")
        capacity_factors = [entry['capacity_factor'] for entry in lcoe_data]
        total_lcoe_values = [entry['categories']['Total LCOE'] for entry in lcoe_data]
        plt.plot(capacity_factors, total_lcoe_values, marker='o', label=key)

    chart_title = f'LCOE vs. Capacity Factor for {key}'
    plt.xlabel('Capacity Factor')
    plt.ylabel('Total LCOE ($/MWh)')
    plt.title('LCOE vs. Capacity Factor')
    plt.legend()
    plt.grid(True)
    plt.xticks(capacity_factors, [f"{cf * 100:.0f}%" for cf in capacity_factors])  # Format capacity factors as percentages
    # Format y-axis labels as currency
    plt.gca().yaxis.set_major_formatter(FuncFormatter(currency_format))

    # Show the plot
    filename = chart_title
    display_figure(fig, f"{filename}.png")
#########################################
def plot_waterfall_chart(lcoe_data_by_generator):
    for key, lcoe_data in lcoe_data_by_generator.items():
        target_capacity_factor = lcoe_data[0]['capacity_factor']
        print(f"\nLCOE Data for {key} at {target_capacity_factor * 100:.1f}% Capacity Factor:\n")
        
        # Assuming you want to plot the first entry (for a specific capacity factor)
        categories = lcoe_data[0]['categories']
        categories_names = list(categories.keys())
        values = list(categories.values())
        
        # Adding total bar
        #total_value = sum(values)
        # Exclude 'Total LCOE' from summation
        if 'Total LCOE' in categories:
            total_lcoe_value = categories['Total LCOE']
            categories_names.remove('Total LCOE')
            values.remove(total_lcoe_value)
        else:
            total_lcoe_value = sum(values)  # Fallback if 'Total LCOE' is not present
        
        
        categories_names.append("Total")
        #values.append(total_value)
        values.append(total_lcoe_value)
        measure = ["relative"] * (len(values) - 1) + ["total"]

        # Formatting the text for each bar with dollar sign and two decimal places
        formatted_values = [f"${v:,.2f}" for v in values]
        
        # Create a Plotly figure
        fig = go.Figure(go.Waterfall(
            name=key, orientation="v",
            measure=measure,
            x=categories_names,
            textposition="outside",
            text=formatted_values[:-1] + [""],  # Removing the default label for the last bar
            y=values,
            increasing={"marker":{"color":"green"}},
            decreasing={"marker":{"color":"red"}},
            totals={"marker":{"color":"blue"}},
            connector={"line":{"color":"rgb(63, 63, 63)"}},
        ))

        # Bold X/Y axis labels using a bold font family
        fig.update_xaxes(title_text="<b>Categories</b>")  # Bold X-axis title
        fig.update_yaxes(title_text="<b>$/MWh</b>")       # Bold Y-axis title

        # Adding an annotation for the last label (total) with a bold font family
        fig.add_annotation(
            x=categories_names[-1],  # Position at the last category
            y=values[-1],            # Position at the value of the total
            text=f"<b>${values[-1]:,.2f}</b>",  # Bold text with formatted value
            showarrow=False,
            yshift=10,  # Adjust this for proper positioning
        )

        # Remove default text for the last bar to prevent overlap
        formatted_values[-1] = ""

        chart_title = f"LCOE for {target_capacity_factor * 100:.1f} percent Capacity Factor for {key}"

        # Update layout for formatting and y-axis range
        fig.update_layout(
            title=dict(
                text = f"Levelized Cost of Energy (LCOE) for {target_capacity_factor * 100:.1f}% Capacity Factor for {key}", 
                font=dict(size=24, color='black'),
                x=0.5, 
                xanchor='center'
            ),
            xaxis_title="<b>Categories</b>",  # Bold X-axis title
            yaxis_title="<b>$/MWh</b>",       # Bold Y-axis title
            yaxis=dict(
                tickprefix="$",
                range=[0, max(values) * 1.2]
            ),
            showlegend=False,
            plot_bgcolor='white'
        )

        # Show the plot
        #filename = chart_title
        filename = sanitize_filename(chart_title)

        #not working for some reason with plotly
        #display_figure(fig, f" {filename}.png")
        fig.show()
#########################################       
# def print_lcoe_table(lcoe_data_by_generator):
#     for key, lcoe_data in lcoe_data_by_generator.items():
#         print(f"\nLCOE Data for {key}:\n")
        
#         # Prepare the table data
#         headers = ["Component"] + [f"{entry['capacity_factor']:.0%}" for entry in lcoe_data]
#         components = lcoe_data[0]['categories'].keys()
#         table = []

#         for component in components:
#             row = [component] + [f"${entry['categories'][component]:,.2f}" for entry in lcoe_data]
#             table.append(row)

#         # Print the table
#         print(key)
#         print(tabulate(table, headers=headers, tablefmt="grid"))

#         # # Save to CSV if a filename is provided
#         # if csv_filename:
#         #     save_table_to_csv(key, headers, table, csv_filename)
#         return key, headers, table
 ########################################   
def print_lcoe_table(lcoe_data_by_generator):
    # Ensure the input is a dictionary with at least one key-value pair
    if not lcoe_data_by_generator:
        print("No data available.")
        return None, None, None
    
    for key, lcoe_data in lcoe_data_by_generator.items():
        print(f"\nLCOE Data for {key}:\n")
        
        # Prepare the table data
        headers = ["Component"] + [f"{entry['capacity_factor']:.0%}" for entry in lcoe_data]
        components = lcoe_data[0]['categories'].keys()
        table = []

        for component in components:
            row = [component] + [f"${entry['categories'][component]:,.2f}" for entry in lcoe_data]
            table.append(row)

        # Print the table
        print(key)
        print(tabulate(table, headers=headers, tablefmt="grid"))

        # Return data for further processing
        return key, headers, table
########################################
def save_table_to_csv(key, headers, table, csv_filename):
    # Create/append to the CSV file
    with open(csv_filename, 'a', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        
        # Write the key as a header
        csvwriter.writerow([f"LCOE Data for {key}"])
        
        # Write the headers
        csvwriter.writerow(headers)
        
        # Write the table data
        for row in table:
            csvwriter.writerow(row)
        
        # Add an empty row for separation
        csvwriter.writerow([])

######################################
# def print_lcoe_table(lcoe_data_by_generator, csv_filename=None):

#     for key, lcoe_data in lcoe_data_by_generator.items():
#         print(f"\nLCOE Data for {key}:\n")
        
#         # Prepare the table data
#         headers = ["Component"] + [f"{entry['capacity_factor']:.0%}" for entry in lcoe_data]
#         components = lcoe_data[0]['categories'].keys()
#         table = []

#         for component in components:
#             row = [component] + [f"${entry['categories'][component]:,.2f}" for entry in lcoe_data]
#             table.append(row)

#         # Print the table
#         print(key)
#         print(tabulate(table, headers=headers, tablefmt="grid"))

#         # Save to CSV if a filename is provided
#         if csv_filename:
#             save_table_to_csv(key, headers, table, csv_filename)
###########################################
# def csv_save_function(df, output_file_path, output_file):
#     if not summary_generator_bid_df.empty:
#         try:
#             df.to_csv(output_file_path + output_file, index=False)
#             print("Data consolidated and saved to output file.")
#         except PermissionError:
#             print(f"Permission denied: Unable to write to the file {output_file}. Please close any application that might be using the file or check your write permissions.")
#         except Exception as e:
#             print(f"An error occurred while saving the file: {e}")
#         else:
#             print("No data to save.")
    
#     return

def show_visualizations_func():
    if cfg.show_visualizatons:
        plt.show()
    else:
        plt.close()
    return

def save_plot(plot, project_name, image_output_path):

    dir_path  = image_output_path

    # Check if the directory exists, if not, create it.
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    # Full path to the file
    full_file_path = os.path.join(dir_path, '{}.png'.format(project_name))

    # Save the figure to the specified directory
    plot.savefig(full_file_path)
    print("Plot saved as '{}'".format(full_file_path))

# Function to remove folder contents at start-up and create csv and image folders to save data in.
def remove_folder_contents(folder):
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                remove_folder_contents(file_path)
                os.rmdir(file_path)
        except Exception as e:
            print(e)# Function to delete files

# Function to delete files
def delete_files(file_list, input_directory):
    for file in file_list:
        file_path = os.path.join(input_directory, file)
        file_path = os.path.normpath(file_path)  # Normalize the path
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Deleted individual file: {file_path}")
            else:
                print(f"File does not exist: {file_path}")
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")

# def delete_folder(folder):
#     if os.path.exists(folder):
#         shutil.rmtree(folder)
#         print(f"Folder and its contents deleted: {folder}")

def delete_folder(folder, retries=3, delay=5):
    for i in range(retries):
        try:
            shutil.rmtree(folder)
            print(f"Successfully deleted the folder: {folder}")
            return
        except PermissionError as e:
            print(f"Attempt {i+1} of {retries} failed: {e}")
            if i < retries - 1:
                time.sleep(delay)
            else:
                print(f"Failed to delete the folder after {retries} attempts.")
                raise

def create_folder(folder):
    try:
        os.makedirs(folder, exist_ok=True)
        print(f"Folder created successfully at: {folder}")
    except Exception as e:
        print(f"An error occurred: {e}")

def consolidate_annual_files(
                forecast_folder, 
                output_data_path, 
                input_data_path, 
                filename_template, 
                datetime_column, 
                output_filename_template
                ):

    global gbl_output_template

    # Define the directory containing the CSV files
    input_directory = input_data_path
    output_directory = output_data_path
    
    # Create an empty list to store individual DataFrames
    data_frames = []

    # List all files in the directory for debugging
    all_files = os.listdir(input_directory)
    print(f"All files in directory: {all_files}")
    print(f"Input directory path: {input_data_path}")

    # Use glob to match the pattern of the files
    search_path = os.path.join(input_directory, filename_template)
    print(f"Searching for files in: {search_path}")
    csv_files = glob.glob(os.path.join(input_directory, filename_template))
    print(f"Files found: {csv_files}")

    # If no files found, raise an error or handle it as needed
    if not csv_files:
        raise FileNotFoundError(f"No files match the pattern {filename_template} in directory {output_data_path}")

    # Extract years from filenames
    years = []

    # Loop through the matched CSV files
    for file in tqdm(csv_files):
        
        # Extract filename from the full path
        filename = os.path.basename(file)

        # Extract year from filename using regex
        match = re.search(r'(\d{4})', filename)
        print(f"Match: {match}")
        if match:
            year = int(match.group(1))
            years.append(year)
        
        # Read each CSV file into a DataFrame
        df = pd.read_csv(file)
        
        # Convert 'begin_datetime_mpt' to datetime
        df[datetime_column] = pd.to_datetime(df[datetime_column])
        
        # Append the DataFrame to the list
        data_frames.append(df)

    # Concatenate all DataFrames in the list
    merged_df = pd.concat(data_frames, ignore_index=True)

    # Optionally, set 'begin_datetime_mpt' as the index
    merged_df.set_index(datetime_column, inplace=True, drop = False)

    # Get the start and end years
    start_year = min(years)
    end_year = max(years)

    path = output_directory
    #Pull the template file which should look like this: "combined_forecast_{}_with_{}.csv"
    file_name_str = gbl.gbl_output_template['combined_forecast_with_Top_Percent_Price_filename_str']
    #output_file_var = f'{output_filename_template}{start_year}_to_{end_year}.csv'
    output_file_var = file_name_str.format(year, top_pct_band.replace(" ", "_"))
    output_file = os.path.join(path, output_file_var)
    merged_df.to_csv(output_file)
    print(f"Consolidated data saved to {path}")

    # Delete the individual files
    delete_files(all_files, input_directory)
    # for file in all_files:
    #     file_path = os.path.join(input_directory, file)
    #     os.remove(file_path)
    #     print(f"Deleted individual files: {file_path}")

    return output_file_var

def replace_underscores_with_spaces_in_list(text_list):
    """
    This function takes a list of strings as input and replaces all underscores with spaces in each string.
    
    :param text_list: list - The list of strings containing underscores
    :return: list - The modified list with spaces instead of underscores in each string

    # Example usage:
    annual_stats_all_years = [
    ['YEAR', 'RECEIVED_POOL_PRICE', 'AVG_POOL_PRICE', 'RECEIVED_POOL_PRICE_RATIO_TO_AVG_SPOT', 
     'MW_PRODUCTION', 'CAPACITY_FACTOR', 'STARTS', 'COLD_STARTS', 'WARM_STARTS', 'HOT_STARTS', 
     'NONE_STARTS', 'STOPS', 'START_COST', 'RUN_HOURS', 'RUN_RATE', 'REVENUE', 'TOTAL_COST_DOLLARS', 
     'OPERATING_MARGIN']]

    # Apply the function to each sublist in the list of lists
    annual_stats_all_years = [replace_underscores_with_spaces_in_list(sublist) for sublist in annual_stats_all_years]


    """
    return [text.replace('_', ' ') for text in text_list]

def save_dataframe_to_csv(df, filename, index_var=True):
    if init.init_ide_option == 'vscode':
        
        output_directory = os.path.join(init.init_base_output_directory, init.init_csv_folder_name_str)
        
        full_path = create_path(output_directory, filename)
        print(f"print file path: {full_path}")
        
        # Ensure the directory exists, not the file itself
        directory = os.path.dirname(full_path)  # Get the directory path
        if not os.path.exists(directory):       # Check if directory exists
            os.makedirs(directory)              # Create the directory if it does not exist

        # New code to handle permission error
        try:
            # Attempt to change permissions only if the file already exists
            if os.path.exists(full_path):
                os.chmod(full_path, 0o666)
                print("File permissions modified successfully!") #new
            else:
                print("File not found:", full_path) #new

            # Attempt to save the DataFrame
            df.to_csv(full_path, index=index_var, header=True)
            print(f"{filename} saved to output folder...") 
            
        except PermissionError:
            print("Permission denied: You don't have the necessary permissions to change the permissions of this file.")
            return None
        except OSError as e:
            print(f"OS error occurred: {e}")
            return None
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return None

    elif init.init_ide_option == 'jupyter_notebook':
        # Implement behavior for Jupyter Notebook if needed
        pass

    elif init.init_ide_option == 'kaggle':
        full_path = create_path(init.init_base_output_directory, filename)
        print(f"print file path: {full_path}")
        
        # Ensure the directory exists, not the file itself
        directory = os.path.dirname(full_path)  # Get the directory path
        if not os.path.exists(directory):       # Check if directory exists
            os.makedirs(directory)              # Create the directory if it does not exist

        try:
            df.to_csv(full_path, index=index_var)  # Save the DataFrame to CSV
            print(f"{filename} saved to output folder...")
        except Exception as e:
            print(f"An unexpected error occurred while saving the file: {e}")
            return None
        
    return
#############################  
def print_worksheet_data(sheet):
    for row in sheet.iter_rows(values_only=True):
        print(row)
#############################   
# Function to clear existing data in a worksheet
def clear_worksheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

#############################
def save_dataframe_to_excel_template(df, output_path, template_path, output_filename, worksheet_name, new_worksheet_name, start_cell='A1'):

    #output_directory = os.path.join(init.gbl_base_output_directory_global, init.gbl_excel_formatted_folder)
    output_full_path = create_path(output_path, output_filename)
    print(f"print file path: {output_full_path}")

    # Load the Excel template
    workbook = load_workbook(template_path)
    print(f" template_path: {template_path}")
    print(f" worksheet_name: {worksheet_name}")
    print(f" workbook.sheetnames: {workbook.sheetnames}")
    
    # Check if the worksheet name exists
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{worksheet_name}' not found in the template.")
    
    sheet = workbook[worksheet_name]

    # Print worksheet data before clearing
    print("Worksheet data before clearing:")
    print_worksheet_data(sheet)

    # Clear existing data in the worksheet
    clear_worksheet(sheet)

    # Print worksheet data after clearing
    print("Worksheet data after clearing:")
    print_worksheet_data(sheet)

    # Write DataFrame to the Excel sheet starting from the specified cell
    start_row = int(start_cell[1:])  # Extract start row from start_cell (e.g., 'A1' -> 1)
    start_col = start_cell[:1]       # Extract start column from start_cell (e.g., 'A1' -> 'A')

    for row in sheet.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None

    # Write DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=ord(start_col) - ord('A') + 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Rename the worksheet
    sheet.title = new_worksheet_name

    # Ensure the directory exists, not the file itself
    directory = os.path.dirname(output_full_path)  # Get the directory path
    if not os.path.exists(directory):       # Check if directory exists
        os.makedirs(directory)              # Create the directory if it does not exist

    # New code to handle permission error
    try:
        # Attempt to change permissions only if the file already exists
        if os.path.exists(output_full_path):
            os.chmod(output_full_path, 0o666)
            print("File permissions modified successfully!") #new
        else:
            print("File not found:", output_full_path) #new

        # Attempt to save the Excel Templplate
        #df.to_csv(full_path, index=index_var, header=True)
        workbook.save(output_full_path)
        print(f"{output_filename} saved to output folder...") 
    
    except PermissionError:
        print("Permission denied: You don't have the necessary permissions to change the permissions of this file.")
        return None
    except OSError as e:
        print(f"OS error occurred: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None